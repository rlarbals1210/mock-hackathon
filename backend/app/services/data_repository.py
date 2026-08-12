from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from statistics import median
from threading import RLock
from typing import Any


KST = timezone(timedelta(hours=9))


@dataclass(frozen=True)
class RouteRecord:
    route_id: str
    origin: str
    origin_region: str
    destination: str
    destination_region: str
    distance_km: int
    standard_hours: float
    toll_krw: int


@dataclass(frozen=True)
class CarrierRecord:
    carrier_id: str
    garage_region: str
    tonnage: int
    body_type: str
    activity_radius: int
    availability_phase: float
    night_allowed: bool
    reliability: float
    preferred_region: str
    historical_acceptance_rate: float = 0.5
    historical_empty_km: int = 25


@dataclass(frozen=True)
class CallRecord:
    call_id: str
    shipper_id: str
    route_id: str
    loading_at: datetime
    registered_at: datetime
    lead_time_hours: float
    loading_window_minutes: int
    tonnage: int
    body_type: str
    item: str
    weight_kg: int
    pallets: int
    base_fare_krw: int
    offered_fare_krw: int
    vehicle_flexible: bool
    split_allowed: bool
    concurrent_load_allowed: bool
    waypoint_allowed: bool
    order_change_allowed: bool
    permission_unapproved: bool
    registration_actor: str
    urgent: bool


@dataclass(frozen=True)
class ShipperStats:
    order_count: int = 0
    weekday_concentration: float = 0.0
    route_concentration: float = 0.0
    segment: str = "스팟"
    historical_acceptance_rate: float = 0.142


@dataclass(frozen=True)
class DataSnapshot:
    routes: dict[str, RouteRecord]
    base_fares: dict[tuple[str, int], int]
    carriers: tuple[CarrierRecord, ...]
    calls: tuple[CallRecord, ...]
    shipper_stats: dict[str, ShipperStats]
    source: str
    warnings: tuple[str, ...] = field(default_factory=tuple)


class DataNotReadyError(RuntimeError):
    pass


def _as_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        result = value
    elif value:
        result = datetime.fromisoformat(str(value))
    else:
        result = datetime.now(KST)
    return result.replace(tzinfo=KST) if result.tzinfo is None else result


def _as_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _yes(value: Any) -> bool:
    return str(value).strip().upper() in {"Y", "TRUE", "1", "예", "허용", "대체허용"}


def _sheet_rows(workbook: Any, name: str) -> list[dict[str, Any]]:
    sheet = workbook[name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    except StopIteration:
        return []
    return [dict(zip(headers, values, strict=False)) for values in iterator if any(value is not None for value in values)]


class MatchingDataRepository:
    def __init__(self, workbook_path: Path, allow_demo_data: bool = True) -> None:
        self.workbook_path = workbook_path
        self.allow_demo_data = allow_demo_data
        self._snapshot: DataSnapshot | None = None
        self._mtime_ns: int | None = None
        self._lock = RLock()

    def snapshot(self) -> DataSnapshot:
        with self._lock:
            current_mtime = self.workbook_path.stat().st_mtime_ns if self.workbook_path.exists() else None
            if self._snapshot is not None and self._mtime_ns == current_mtime:
                return self._snapshot
            if self.workbook_path.exists():
                self._snapshot = self._load_workbook(self.workbook_path)
            elif self.allow_demo_data:
                self._snapshot = self._demo_snapshot()
            else:
                raise DataNotReadyError(f"매칭 데이터 파일을 찾을 수 없습니다: {self.workbook_path}")
            self._mtime_ns = current_mtime
            return self._snapshot

    def _load_workbook(self, path: Path) -> DataSnapshot:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DataNotReadyError("엑셀 데이터를 읽기 위한 openpyxl이 설치되지 않았습니다.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        required = {"콜등록이력", "운송인마스터", "운송인이력", "참조_노선", "참조_기준운임"}
        missing = sorted(required.difference(workbook.sheetnames))
        if missing:
            raise DataNotReadyError(f"필수 시트가 없습니다: {', '.join(missing)}")

        route_rows = _sheet_rows(workbook, "참조_노선")
        fare_rows = _sheet_rows(workbook, "참조_기준운임")
        carrier_rows = _sheet_rows(workbook, "운송인마스터")
        history_rows = _sheet_rows(workbook, "운송인이력")
        call_rows = _sheet_rows(workbook, "콜등록이력")
        proposal_rows = _sheet_rows(workbook, "개입제안로그") if "개입제안로그" in workbook.sheetnames else []
        workbook.close()

        routes = {
            str(row["노선ID"]): RouteRecord(
                route_id=str(row["노선ID"]),
                origin=str(row["출발지"]),
                origin_region=str(row["출발권역"]),
                destination=str(row["도착지"]),
                destination_region=str(row["도착권역"]),
                distance_km=_as_int(row["거리km"]),
                standard_hours=_as_float(row["표준소요_h"]),
                toll_krw=_as_int(row["톨비"]),
            )
            for row in route_rows
        }
        base_fares = {
            (str(row["노선ID"]), _as_int(row["톤급"])): _as_int(row["기준운임"])
            for row in fare_rows
        }

        accept_counts: dict[str, list[int]] = defaultdict(list)
        empty_distances: dict[str, list[int]] = defaultdict(list)
        for row in history_rows:
            carrier_id = str(row.get("차주ID", ""))
            if not carrier_id:
                continue
            accept_counts[carrier_id].append(1 if str(row.get("수락여부", "")) == "수락" else 0)
            empty_distances[carrier_id].append(_as_int(row.get("공차거리_km"), 25))

        carriers = []
        for row in carrier_rows:
            carrier_id = str(row["차주ID"])
            accepts = accept_counts.get(carrier_id, [])
            empties = empty_distances.get(carrier_id, [])
            carriers.append(CarrierRecord(
                carrier_id=carrier_id,
                garage_region=str(row["차고지"]),
                tonnage=_as_int(row["톤급"]),
                body_type=str(row["적재형태"]),
                activity_radius=_as_int(row["활동반경"], 1),
                availability_phase=_as_float(row["가용위상"], 1.0),
                night_allowed=bool(row["야간수용"]),
                reliability=_as_float(row["이행률"], 0.8),
                preferred_region=str(row["명시선호권역"]),
                historical_acceptance_rate=(sum(accepts) + 2.0) / (len(accepts) + 4.0),
                historical_empty_km=max(1, int(median(empties))) if empties else 25,
            ))

        calls = []
        for row in call_rows:
            registered_at = _as_datetime(row["등록일시"])
            loading_at = _as_datetime(row["상차희망"])
            calls.append(CallRecord(
                call_id=str(row["콜ID"]),
                shipper_id=str(row["화주ID"]),
                route_id=str(row["노선ID"]),
                loading_at=loading_at,
                registered_at=registered_at,
                lead_time_hours=_as_float(row["리드타임_h"]),
                loading_window_minutes=_as_int(row["시간창_분"]),
                tonnage=_as_int(row["톤급"]),
                body_type=str(row["적재형태"]),
                item=str(row["품목"]),
                weight_kg=_as_int(row["중량_kg"]),
                pallets=_as_int(row["파렛트"]),
                base_fare_krw=_as_int(row["기준운임"]),
                offered_fare_krw=_as_int(row["제시운임"]),
                vehicle_flexible=str(row.get("차량유연성", "")) == "대체허용",
                split_allowed=_yes(row.get("분할운송")),
                concurrent_load_allowed=_yes(row.get("동시적재")),
                waypoint_allowed=_yes(row.get("경유허용")),
                order_change_allowed=_yes(row.get("상하차_순서변경")),
                permission_unapproved=str(row.get("원화주_조정권한_증빙", "")) == "미승인",
                registration_actor=str(row.get("등록주체", "원화주직접")),
                urgent=str(row.get("긴급여부", "일반")) == "긴급",
            ))

        shipper_stats = self._build_shipper_stats(calls, proposal_rows)
        return DataSnapshot(
            routes=routes,
            base_fares=base_fares,
            carriers=tuple(carriers),
            calls=tuple(calls),
            shipper_stats=shipper_stats,
            source=path.name,
        )

    @staticmethod
    def _build_shipper_stats(calls: list[CallRecord], proposal_rows: list[dict[str, Any]]) -> dict[str, ShipperStats]:
        grouped: dict[str, list[CallRecord]] = defaultdict(list)
        call_to_shipper = {}
        for call in calls:
            grouped[call.shipper_id].append(call)
            call_to_shipper[call.call_id] = call.shipper_id

        reactions: dict[str, list[int]] = defaultdict(list)
        for row in proposal_rows:
            shipper_id = call_to_shipper.get(str(row.get("콜ID", "")))
            if shipper_id:
                reactions[shipper_id].append(1 if str(row.get("사용자반응", "")) == "수락" else 0)
        global_reactions = [value for values in reactions.values() for value in values]
        global_rate = sum(global_reactions) / len(global_reactions) if global_reactions else 0.142

        result = {}
        for shipper_id, items in grouped.items():
            ordered = sorted(items, key=lambda item: item.registered_at)
            weekdays = Counter(item.loading_at.weekday() for item in ordered)
            routes = Counter(item.route_id for item in ordered)
            dates = [item.registered_at.date() for item in ordered]
            gaps = [(right - left).days for left, right in zip(dates, dates[1:], strict=False)]
            active_weeks = sorted({date - timedelta(days=date.weekday()) for date in dates})
            week_gaps = [(right - left).days for left, right in zip(active_weeks, active_weeks[1:], strict=False)]
            weekday_concentration = max(weekdays.values()) / len(ordered)
            route_concentration = max(routes.values()) / len(ordered)
            mean_gap = sum(gaps) / len(gaps) if gaps else 999.0
            median_week_gap = median(week_gaps) if week_gaps else 999.0
            if len(ordered) >= 3 and weekday_concentration >= 0.6 and 6 <= median_week_gap <= 8:
                segment = "고정주간"
            elif len(ordered) >= 8 and mean_gap <= 5.5:
                segment = "고빈도"
            elif len(ordered) >= 3 and 12.5 <= median_week_gap <= 15.5:
                segment = "격주"
            else:
                segment = "스팟"
            shipper_reactions = reactions.get(shipper_id, [])
            smoothed_rate = (sum(shipper_reactions) + 10 * global_rate) / (len(shipper_reactions) + 10)
            result[shipper_id] = ShipperStats(
                order_count=len(ordered),
                weekday_concentration=weekday_concentration,
                route_concentration=route_concentration,
                segment=segment,
                historical_acceptance_rate=smoothed_rate,
            )
        return result

    @staticmethod
    def _demo_snapshot() -> DataSnapshot:
        routes = {
            "R01": RouteRecord("R01", "부산신항", "영남", "김포", "수도권", 400, 6.5, 32800),
            "R04": RouteRecord("R04", "창원공단", "영남", "평택", "수도권", 300, 5.2, 24600),
            "R12": RouteRecord("R12", "대전유성", "충청", "김해", "영남", 200, 3.9, 16400),
        }
        base_fares = {
            ("R01", 5): 368000, ("R01", 11): 454000, ("R01", 25): 596000,
            ("R04", 5): 326000, ("R04", 11): 406000, ("R04", 25): 520000,
            ("R12", 5): 281000, ("R12", 11): 350000, ("R12", 25): 439000,
        }
        def pseudo(index: int, salt: int) -> float:
            value = (index + salt * 0x9E3779B9) & 0xFFFFFFFF
            value ^= value >> 16
            value = (value * 0x7FEB352D) & 0xFFFFFFFF
            value ^= value >> 15
            value = (value * 0x846CA68B) & 0xFFFFFFFF
            value ^= value >> 16
            return value / 2**32

        def weighted(value: float, options: tuple[tuple[Any, float], ...]) -> Any:
            cumulative = 0.0
            for item, probability in options:
                cumulative += probability
                if value < cumulative:
                    return item
            return options[-1][0]

        garage_options = (("영남", .30), ("수도권", .33), ("충청", .14), ("호남", .13), ("강원제주", .10))
        tonnage_options = ((5, .50), (11, .33), (25, .17))
        body_options = (("냉동", .18), ("냉장", .02), ("윙바디", .29), ("카고", .335), ("탑차", .175))
        radius_options = ((1, .40), (2, .40), (3, .20))
        preferred_options = (("영남", .30), ("수도권", .27), ("충청", .17), ("호남", .15), ("강원제주", .11))
        carriers = tuple(
            CarrierRecord(
                carrier_id=f"D{index:05d}",
                garage_region=weighted(pseudo(index, 1), garage_options),
                tonnage=weighted(pseudo(index, 2), tonnage_options),
                body_type=weighted(pseudo(index, 3), body_options),
                activity_radius=weighted(pseudo(index, 4), radius_options),
                availability_phase=pseudo(index, 5),
                night_allowed=pseudo(index, 6) < .42,
                reliability=round(.60 + pseudo(index, 7) * .399, 3),
                preferred_region=weighted(pseudo(index, 8), preferred_options),
                historical_acceptance_rate=.25 + pseudo(index, 9) * .65,
                historical_empty_km=1 + int(pseudo(index, 10) * 89),
            )
            for index in range(1, 11_801)
        )
        calls = (
            CallRecord("C0042", "S018", "R01", datetime(2026, 8, 13, 17, 30, tzinfo=KST), datetime(2026, 8, 12, 15, 30, tzinfo=KST), 26, 180, 11, "카고", "철강재", 9500, 10, 454000, 454000, False, False, False, False, False, False, "원화주직접", False),
            CallRecord("C0178", "S044", "R12", datetime(2026, 8, 13, 19, 0, tzinfo=KST), datetime(2026, 8, 12, 12, 0, tzinfo=KST), 31, 240, 11, "카고", "생활용품", 7800, 12, 350000, 350000, True, False, True, False, False, False, "원화주직접", False),
            CallRecord("C0224", "S087", "R04", datetime(2026, 8, 14, 9, 0, tzinfo=KST), datetime(2026, 8, 12, 10, 0, tzinfo=KST), 47, 480, 11, "카고", "자동차부품", 8200, 13, 406000, 406000, True, False, False, True, False, False, "원화주직접", False),
        )
        stats = {
            "S018": ShipperStats(18, 0.72, 0.81, "고정주간", 0.22),
            "S044": ShipperStats(12, 0.50, 0.58, "고빈도", 0.18),
            "S087": ShipperStats(7, 0.43, 0.43, "스팟", 0.14),
        }
        return DataSnapshot(
            routes=routes,
            base_fares=base_fares,
            carriers=carriers,
            calls=calls,
            shipper_stats=stats,
            source="demo_seed_v1",
            warnings=("생성 엑셀 파일이 없어 로컬 연동용 데모 데이터를 사용합니다.",),
        )
